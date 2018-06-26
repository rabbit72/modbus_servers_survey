import os
import win_inet_pton
from datetime import datetime
from pyModbusTCP.client import ModbusClient
from openpyxl import Workbook
from socket import gaierror

STATUS_NAMES = [
    'Tmax_on_com',
    'Tmim_on_com',
    'FIRE-r',
    'HOT_1_ON_r',
    'HOT_2_ON_r',
    'WENT',
    'COND_1_ON_r',
    'COND_2_ON_r',
    'ALARM_SRK',
    'Door_on_r',
    'ISP1',
    'ISP2',
    'COND_1_ALARM',
    'COND_2_ALARM'
]
VALUE_NAMES = [
    'AC1',
    'AC2',
    'Up',
    'Down'
]


def get_list_servers():
    list_servers = []
    with open('hosts.txt', 'r') as file_with_servers:
        for server in file_with_servers:
            host, name_server = server.rstrip().split(',')
            list_servers.append((host, name_server))
    return list_servers


def save_data_xlsx(servers_data):
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'log_servers'
    columns_names = []
    columns_names.extend(STATUS_NAMES)
    columns_names.extend(VALUE_NAMES)
    row_names = ['Name server']
    row_names.extend(columns_names)
    sheet.append(row_names)

    row = 2
    for server_name, data_server in servers_data.items():
        col = 1
        sheet.cell(row=row, column=col).value = server_name
        col += 1
        if not data_server:
            sheet.cell(row=row, column=col).value = 'No connection'
            row += 1
            continue
        for value_name in columns_names:
            sheet.cell(row=row, column=col).value = data_server[value_name]
            col += 1
        row += 1

    time_point = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    dir_for_file = './logs/'
    if not os.path.exists(dir_for_file):
        os.makedirs(dir_for_file)
    name_file = os.path.join(dir_for_file, '{0}.xlsx'.format(time_point))
    wb.save(name_file)


def read_server(host):
    registers = None
    start_register = 0
    number_of_registers = 5
    try:
        client = ModbusClient(host=host, timeout=1)
        client.open()
        if client.is_open():
            registers = client.read_holding_registers(
                start_register,
                number_of_registers
            )
    except gaierror:
        pass
    finally:
        return registers


def processing_equipment_status(register):
    number_of_bin = 16
    cut_size = number_of_bin - len(STATUS_NAMES)
    bin_status = format(register, 'b').zfill(number_of_bin)
    status_cut = bin_status[cut_size:]
    inverted_status = status_cut[::-1]
    equipment_status = {}
    for key, status in zip(STATUS_NAMES, inverted_status):
        equipment_status[key] = bool(int(status))
    return equipment_status


def get_scaling_temperature_values(values):
    """
    Temperature matching:
    0 is 0 C and 32768 is 3276.8 degree C
    65535 is -1 degree C and 32769 is -3276.7 degree C
    """
    temperature_values = {}
    for key, value in zip(VALUE_NAMES, values):
        if value > 32768:
            value -= 65536
        temperature_values[key] = value / 10
    return temperature_values


def get_data_servers(servers):
    data_servers = {}
    for host, name_server in servers:
        response = read_server(host)
        if not response:
            data_servers[name_server] = response
            continue
        raw_status, raw_temp_values = response[0], response[1:]
        status = processing_equipment_status(raw_status)
        temp_values = get_scaling_temperature_values(raw_temp_values)
        data_servers[name_server] = status.copy()
        data_servers[name_server].update(temp_values)
    return data_servers


if __name__ == '__main__':
    list_servers = get_list_servers()
    data_servers = get_data_servers(list_servers)
    save_data_xlsx(data_servers)
